"""
report_export.py
=================

Builds a downloadable Excel (.xlsx) report from an already-computed
`KineticsResult`, using `openpyxl`. This module performs NO kinetics
calculations of its own - every number it writes was already produced
by `core.calculations.analyze_kinetics`. It only formats those numbers
into a workbook with four sheets, as requested for the final
packaging step:

    1. Input Data       - the (time, concentration) pairs that were
                           actually fed into the analysis.
    2. Processed Data   - the per-point results (model-predicted
                           concentration, residuals, conversion, rate)
                           already shown in the app's detail table.
    3. Kinetic Results  - the fitted parameters (n, k, R^2, half-life,
                           time for complete conversion, etc.).
    4. Final Results    - a short summary (final concentration/
                           conversion/rate) plus the optional AI
                           interpretation text, if one was generated.

Kept separate from `streamlit_app.py` so it can be unit tested without
importing Streamlit.
"""

import io
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from core.calculations import KineticsResult

_HEADER_FONT = Font(bold=True)
_TITLE_FONT = Font(bold=True, size=14)


def _autosize(ws, min_width=10, max_width=40):
    """Best-effort column widths based on cell content length."""
    for col_cells in ws.columns:
        length = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 2, min_width), max_width)


def _write_table(ws, headers, rows, start_row):
    """Write a header row followed by data rows, starting at start_row."""
    for j, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=j, value=h).font = _HEADER_FONT
    for i, row in enumerate(rows, start=start_row + 1):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)
    return start_row + len(rows)


def _write_kv(ws, pairs, start_row):
    """Write a list of (label, value) pairs as two columns."""
    for i, (k, v) in enumerate(pairs, start=start_row):
        ws.cell(row=i, column=1, value=k).font = _HEADER_FONT
        ws.cell(row=i, column=2, value=v)
    return start_row + len(pairs)


def build_excel_report(
    result: KineticsResult,
    source_label: Optional[str] = None,
    time_unit: Optional[str] = None,
    concentration_unit: Optional[str] = None,
    ai_interpretation: Optional[str] = None,
) -> bytes:
    """
    Assemble the 4-sheet Excel workbook described above from an
    already-computed `KineticsResult`. Nothing here is recalculated -
    every value is read directly off `result`, or is one of the unit/
    source strings the UI already had.

    Returns
    -------
    bytes
        Raw .xlsx file content, ready to hand to
        `st.download_button(data=...)`.
    """
    time_header = f"time ({time_unit})" if time_unit else "time"
    conc_header = f"concentration ({concentration_unit})" if concentration_unit else "concentration"
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    wb = Workbook()

    # ---- Sheet 1: Input Data --------------------------------------
    ws_input = wb.active
    ws_input.title = "Input Data"
    ws_input.cell(row=1, column=1, value="Reaction Kinetics Analyzer - Input Data").font = _TITLE_FONT
    ws_input.cell(row=2, column=1, value=f"Source: {source_label or 'N/A'}")
    ws_input.cell(row=3, column=1, value=f"Generated: {generated_at}")
    _write_table(
        ws_input,
        [time_header, conc_header],
        list(zip(result.time.tolist(), result.concentration.tolist())),
        start_row=5,
    )
    _autosize(ws_input)

    # ---- Sheet 2: Processed Data (per-point analysis results) -----
    ws_proc = wb.create_sheet("Processed Data")
    ws_proc.cell(row=1, column=1, value="Processed Data (per data point)").font = _TITLE_FONT
    residuals = (result.concentration - result.predicted_concentration).tolist()
    _write_table(
        ws_proc,
        [
            time_header,
            f"experimental {conc_header}",
            f"model-predicted {conc_header}",
            "residual (exp - model)",
            "conversion X_A",
            "rate of reaction (-r_A)",
        ],
        list(zip(
            result.time.tolist(),
            result.concentration.tolist(),
            result.predicted_concentration.tolist(),
            residuals,
            result.conversion.tolist(),
            result.rate_at_each_point.tolist(),
        )),
        start_row=3,
    )
    _autosize(ws_proc)

    # ---- Sheet 3: Kinetic Results -----------------------------------
    ws_kin = wb.create_sheet("Kinetic Results")
    ws_kin.cell(row=1, column=1, value="Kinetic Results").font = _TITLE_FONT
    next_row = _write_kv(
        ws_kin,
        [
            ("Reaction order (n)", result.reaction_order),
            ("Rate constant (k)", result.rate_constant),
            ("Rate constant units", result.rate_constant_units),
            ("R-squared (goodness of fit)", result.r_squared),
            ("Half-life (t_1/2)", result.half_life if result.half_life is not None else "N/A"),
            (
                "Time for complete conversion",
                result.time_for_complete_conversion
                if result.time_for_complete_conversion is not None else "N/A",
            ),
            ("Initial concentration (C_A0)", result.CA0),
        ],
        start_row=3,
    )
    if result.warnings:
        ws_kin.cell(row=next_row + 1, column=1, value="Warnings:").font = _HEADER_FONT
        for i, w in enumerate(result.warnings, start=next_row + 2):
            ws_kin.cell(row=i, column=1, value=w)
    _autosize(ws_kin)

    # ---- Sheet 4: Final Results ------------------------------------
    ws_final = wb.create_sheet("Final Results")
    ws_final.cell(row=1, column=1, value="Final Results Summary").font = _TITLE_FONT
    next_row = _write_kv(
        ws_final,
        [
            ("Final concentration (C_A, last point)", float(result.concentration[-1])),
            ("Final conversion (X_A, last point)", float(result.conversion[-1])),
            ("Rate of reaction at C_A0", result.rate_at_CA0),
            ("Rate of reaction at final point", float(result.rate_at_each_point[-1])),
            ("Number of data points", len(result.time)),
        ],
        start_row=3,
    )
    _autosize(ws_final)

    if ai_interpretation:
        label_row = next_row + 1
        text_row = label_row + 1
        ws_final.cell(row=label_row, column=1, value="AI Interpretation:").font = _HEADER_FONT
        cell = ws_final.cell(row=text_row, column=1, value=ai_interpretation)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws_final.row_dimensions[text_row].height = 120
        ws_final.column_dimensions["A"].width = 100

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
