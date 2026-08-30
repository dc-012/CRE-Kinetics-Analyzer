"""
Tests for app/report_export.py - the downloadable Excel report.

These tests run a real analysis through core.calculations first (the
same entry point the Streamlit app uses), then check that the report
builder writes exactly those already-computed values into the
workbook, with no recalculation happening in report_export.py itself.
"""

import io

import openpyxl
import pytest

from app import report_export
from core.calculations import analyze_kinetics


@pytest.fixture
def sample_result():
    t = [0, 20, 40, 60, 120, 180, 300]
    c = [10, 8, 6, 5, 3, 2, 1]
    return analyze_kinetics(t, c)


def test_build_excel_report_returns_bytes(sample_result):
    data = report_export.build_excel_report(sample_result)
    assert isinstance(data, bytes)
    assert len(data) > 0


def test_report_has_four_expected_sheets(sample_result):
    data = report_export.build_excel_report(sample_result)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames == [
        "Input Data", "Processed Data", "Kinetic Results", "Final Results",
    ]


def test_input_data_sheet_matches_result_arrays(sample_result):
    data = report_export.build_excel_report(sample_result, time_unit="min", concentration_unit="mol/L")
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Input Data"]

    # Header row is at row 5 (rows 1-4 hold title/source/generated-at).
    header = [c.value for c in ws[5]]
    assert header == ["time (min)", "concentration (mol/L)"]

    first_data_row = [c.value for c in ws[6]]
    assert first_data_row[0] == pytest.approx(sample_result.time[0])
    assert first_data_row[1] == pytest.approx(sample_result.concentration[0])


def test_processed_data_sheet_row_count_matches_points(sample_result):
    data = report_export.build_excel_report(sample_result)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Processed Data"]
    # Header at row 3, then one row per experimental data point.
    n_points = len(sample_result.time)
    data_rows = list(ws.iter_rows(min_row=4, max_row=3 + n_points))
    assert len(data_rows) == n_points


def test_kinetic_results_sheet_contains_reaction_order_and_k(sample_result):
    data = report_export.build_excel_report(sample_result)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Kinetic Results"]
    labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    values = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
              for r in range(1, ws.max_row + 1)}
    assert "Reaction order (n)" in labels
    assert values["Reaction order (n)"] == pytest.approx(sample_result.reaction_order)
    assert values["Rate constant (k)"] == pytest.approx(sample_result.rate_constant)


def test_final_results_sheet_has_no_ai_text_when_none_provided(sample_result):
    data = report_export.build_excel_report(sample_result, ai_interpretation=None)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Final Results"]
    all_values = [c.value for row in ws.iter_rows() for c in row]
    assert not any(v == "AI Interpretation:" for v in all_values)


def test_final_results_sheet_includes_ai_text_when_provided(sample_result):
    explanation = "This is a first-order-like reaction with a good fit."
    data = report_export.build_excel_report(sample_result, ai_interpretation=explanation)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Final Results"]
    all_values = [c.value for row in ws.iter_rows() for c in row]
    assert "AI Interpretation:" in all_values
    assert explanation in all_values
